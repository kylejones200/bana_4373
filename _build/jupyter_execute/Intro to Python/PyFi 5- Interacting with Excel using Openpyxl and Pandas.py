#!/usr/bin/env python
# coding: utf-8

# # Interacting with Excel

# For many use cases, you'll only need Pandas, since it contains some of the functionality of the openpyxl package right out of the box.  Other times, you'll want to `import openpyxl` directly.  In this case, we'll explicitly `import openpyxl`, as this allows us to write data to an existing Excel file.
# 
# 
# Openpyxl is a library for manipulating Excel workbooks and worksheets. In this example we read in a workbook, navigate to a specific sheet and save the values on that sheet as a Pandas DataFrame. Then we can use Pandas methods to clean the data and save it back to the same workbook in a new worksheet.
# 
# The advantages of Python in this case are:
# * Speed of processing: Pandas can process calculations faster than Excel
# * Replicability: It is easier to apply these transformations to another dataset using Python
# * Readability: Python is easier to read and understand the various transformations applied to the data
# * Automation: The Python script could be executed automatically to clean the data without human intervention
# 
# ## When to choose Excel or Python
# Excel is good for small, quick analyses. It allows you to easily see the data in a familiar format. It can present data in tables and basic charts. 
# 
# Python out preforms Excel if you are working with lots of data (Excel max row count is 1 million), applying formulas (especially lookups, matches, and sumifs), and using array functions. 
# 
# There are data wrangling features in Python that are not available in Excel such as stack/unstack, unpivot, and melt. 
# 
# If a worksheet relies heavily on VBA, it is probably a good candidate for converting to Python. Doing the analysis/transforms in Python will be faster and avoid manual errors which can easily happen in Excel. 
# 
# # Tutorial Overview
# In this module, we will:
# * import an excel file using Openpyxl
# * remove the excess columns and rows
# * export to Excel

# In[1]:


import pandas as pd
import openpyxl


# ### Specify paths to input and output files

# In[2]:


path_to_data = 'data/excel_input.xlsx'
path_to_output = 'data/excel_output_new_file.xlsx'

print('The path to the Excel file we are going to read is:')
print(path_to_data)

print('\nThe path to the new Excel file we are going to create is:')
print(path_to_output)


# ### Read data from Excel

# In[3]:


retail = pd.read_excel(path_to_data, sheet_name='retail')
mfn = pd.read_excel(path_to_data, sheet_name='mfn')

print('retail data:\n')
print(retail)

print('\nmfn data:\n')
print(mfn)


# ### Generate the desired output
# In this case, we'll create another data frame that will become a tab in Excel

# In[4]:


combined = pd.concat([retail, mfn])

print('combined data:\n')
print(combined)


# ### Option 1: Create a new Excel file

# In[5]:


combined.to_excel(path_to_output, sheet_name='combined', index=False)


# ### Option 2: Write new data to the same Excel file
# 
# WARNING: If used incorrectly, this can cause data loss in your original file.  If you aren't sure, create a backup of your input file first.

# In[6]:


# create a workbook object using openpyxl
workbook = openpyxl.load_workbook(path_to_data)

# create a pandas ExcelWriter object from the same source file
writer = pd.ExcelWriter(path_to_data, engine='openpyxl') 

# link the two objects together
writer.book = workbook

# save
combined.to_excel(writer, sheet_name='combined', index=False)
writer.save()


# # A worked example
# 
# This example cleans a data file. Instructions for doing this cleaning using native Excel are included in the Workbook. 

# In[7]:


import pandas as pd
import openpyxl 

filename = 'data/XL-02-PC-05-Cleaning-Up-Data-Before.xlsx'

wb = openpyxl.load_workbook(filename)
ws =  wb['Data']
df = pd.DataFrame(ws.values)

# Makes first line the column names
def make_header(df: pd.DataFrame)-> pd.DataFrame:
    df.columns = df.iloc[0]
    df.drop(df.index[0], inplace=True)
    return df

df = make_header(df)

# Expanding data in single columns to multiple columns
df[['First Name', 'Last Name']]= df['Customer Name'].str.split(" ", n=1, expand=True)
df[['Address', 'City', 'State', 'ZIP']]= df['Address, City, State, and ZIP'].str.split(",", n=3, expand=True)

# Changing the case of new columns
df['Address'] = df['Address'].str.title()
df['City'] = df['City'].str.title()
df['State'] = df['State'].str.upper()
                                                               
from openpyxl.utils.dataframe import dataframe_to_rows

# Change the file name if you want to preserve the origional version
wb.create_sheet(index=0, title='Cleaned Data') 
ws =  wb['Cleaned Data']

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

wb.save(filename)
wb.close()


# # Extensions
# 
# This section lists some ideas for extending the tutorial that you may wish to explore.
# * Describe three examples when Openpyxl would be better than using Excel directly.
# 
# # Further Reading
# This section provides more resources on the topic if you are looking to go deeper.
# 
# ## Books
# * Python for Data Analysis, by William McKinney. http://shop.oreilly.com/product/0636920023784.do
# * Automate the boring stuff with Python, by Al Sweigart. See Chapter 13. https://automatetheboringstuff.com/
# 
# ## APIs
# * openpyxl-A Python library to read/write Excel 2010 xlsx/xlsm files.  http://openpyxl.readthedocs.org/en/default
# 
# ## Articles
# * Working with Excel sheets in Python using openpyxl. https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f  
# 
# # Summary
# 
# In this tutorial, you worked with Openpyxl. Specifically, you learned:
# * How to import an excel file using Openpyxl
# * How to remove the excess columns and rows
# * How to export to to an excel file

# In[ ]:




